import os
import io
from datetime import datetime, timedelta
from dotenv import load_dotenv
import httpx
import openpyxl

load_dotenv()

# Índices fijos de columnas (fila 1 = encabezados reales, fila 0 = títulos agrupadores)
COL_TIPO       = 0
COL_SUBREGION  = 1
COL_MUNICIPIO  = 2
COL_CNT_ANT    = 3
COL_MES_ANT    = 4
COL_MES_ACT    = 5
COL_ENCARGADO  = 6
COL_CNT_ACT    = 7
COL_FECHA_ELAB = 8
COL_PUBLICADA  = 9

MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


def _share_url_to_download(url: str) -> str:
    """Convierte un link de compartir de SharePoint en URL de descarga directa."""
    # Agregar &download=1 fuerza la descarga del archivo
    separator = "&" if "?" in url else "?"
    return url + separator + "download=1"


def _excel_serial_to_mes_anio(serial) -> tuple[str, str]:
    try:
        n = int(float(str(serial)))
        if n < 1:
            return "", ""
        dt = datetime(1899, 12, 30) + timedelta(days=n)
        return MESES_ES[dt.month - 1], str(dt.year)
    except Exception:
        return str(serial).strip(), ""


def _parse_mes_anio(texto: str) -> tuple[str, str]:
    s = str(texto).strip()
    if not s:
        return "", ""
    try:
        float(s)
        return _excel_serial_to_mes_anio(s)
    except ValueError:
        pass
    partes = s.split()
    if len(partes) >= 2 and partes[-1].isdigit():
        return " ".join(partes[:-1]), partes[-1]
    return s, ""


def _cell(row, idx):
    try:
        v = row[idx]
        return v
    except IndexError:
        return None


def _str(row, idx) -> str:
    v = _cell(row, idx)
    if v is None:
        return ""
    return str(v).strip()


def _int(row, idx) -> int:
    v = _cell(row, idx)
    if v is None or str(v).strip() == "":
        return 0
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return 0


async def fetch_fichas() -> list[dict]:
    share_url = os.getenv("MS_SHARE_URL", "")
    sheet_name = os.getenv("MS_SHEET_NAME", "Seguimiento a Fichas")

    if not share_url:
        raise ValueError("MS_SHARE_URL no está configurado en .env")

    download_url = _share_url_to_download(share_url)

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        resp = await client.get(download_url)
        resp.raise_for_status()

    # Leer el Excel desde memoria con openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {wb.sheetnames}")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Fila 0 = títulos agrupadores, Fila 1 = encabezados → datos desde fila 2
    if len(all_rows) < 3:
        return []

    rows = []
    for raw in all_rows[2:]:
        # Ignorar filas vacías
        if not any(c for c in raw if c not in (None, "", " ")):
            continue

        municipio  = _str(raw, COL_MUNICIPIO)
        subregion  = _str(raw, COL_SUBREGION)
        encargado  = _str(raw, COL_ENCARGADO)
        tipo       = _str(raw, COL_TIPO)
        mes_act    = _str(raw, COL_MES_ACT)
        mes_ant    = _str(raw, COL_MES_ANT)
        publicada  = _str(raw, COL_PUBLICADA)
        cnt_act    = _int(raw, COL_CNT_ACT)
        cnt_ant    = _int(raw, COL_CNT_ANT)

        # Fecha elaboración → mes y año de realización para contrato actual
        fecha_raw = _cell(raw, COL_FECHA_ELAB)
        if isinstance(fecha_raw, datetime):
            # openpyxl ya convierte fechas a datetime
            mes_real_act = MESES_ES[fecha_raw.month - 1]
            anio_real_act = str(fecha_raw.year)
        else:
            mes_real_act, anio_real_act = _parse_mes_anio(str(fecha_raw) if fecha_raw else "")

        # ── Contrato ACTUAL ──────────────────────────────────────────
        if cnt_act > 0 or mes_act:
            rows.append({
                "municipio":        municipio,
                "subregion":        subregion,
                "tipo_ficha":       tipo,
                "encargado":        encargado,
                "mes_asignacion":   mes_act,
                "mes_realizacion":  mes_real_act,
                "anio_realizacion": anio_real_act,
                "contrato":         "este",
                "cantidad_fichas":  cnt_act,
                "publicada":        publicada,
            })

        # ── Contrato ANTERIOR ────────────────────────────────────────
        if cnt_ant > 0 or mes_ant:
            mes_real_ant, anio_real_ant = _parse_mes_anio(mes_ant)
            if not anio_real_ant:
                anio_real_ant = "2025"
            rows.append({
                "municipio":        municipio,
                "subregion":        subregion,
                "tipo_ficha":       tipo,
                "encargado":        encargado,
                "mes_asignacion":   "",
                "mes_realizacion":  mes_real_ant,
                "anio_realizacion": anio_real_ant,
                "contrato":         "anterior",
                "cantidad_fichas":  cnt_ant,
                "publicada":        publicada,
            })

    wb.close()
    return rows
